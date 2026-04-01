from __future__ import annotations

import logging
import os
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from app import models
from app.auth import hash_password
from app.storage import DB


logger = logging.getLogger(__name__)


def _normalize_key(name: str) -> str:
    s = str(name or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    s = re.sub(r"[\s\-/\.'’]+", "_", s)
    s = re.sub(r"[^a-z0-9_]+", "", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _to_int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int,)):
        return int(value)
    if isinstance(value, float):
        return int(round(value))
    s = _to_text(value).replace(",", ".")
    try:
        f = float(s)
        return int(round(f))
    except Exception:
        return 0


_DATE_DMY_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")


def _to_date_iso(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    s = _to_text(value)
    if not s:
        return None

    # ISO date or datetime
    try:
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
            return date.fromisoformat(s).isoformat()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(:\d{2})?", s):
            return datetime.fromisoformat(s.replace(" ", "T")).date().isoformat()
    except Exception:
        pass

    # dd/MM/yyyy
    m = _DATE_DMY_RE.match(s)
    if m:
        dd, mm, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(yyyy, mm, dd).isoformat()
        except ValueError:
            return None

    return None


def _first_present(rec: Dict[str, Any], *keys: str) -> Any:
    for k in keys:
        if k in rec and _to_text(rec.get(k)):
            return rec.get(k)
    norm = {_normalize_key(k): v for k, v in rec.items()}
    for k in keys:
        nk = _normalize_key(k)
        if nk in norm and _to_text(norm.get(nk)):
            return norm.get(nk)
    return None


def _gen_id(prefix: str, *parts: Any, max_len: int = 64) -> str:
    raw = "_".join(_to_text(p) for p in parts if _to_text(p))
    raw = _normalize_key(raw) or "x"
    out = f"{prefix}-{raw}"
    if len(out) <= max_len:
        return out
    # keep deterministic suffix
    return out[: max_len - 9] + "-" + out[-8:]


def _find_header_row(ws, needle: str, *, max_rows: int = 60) -> Optional[Tuple[int, List[str]]]:
    n = needle.lower().strip()
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        for v in row:
            if v is None:
                continue
            if n in str(v).lower():
                headers = [str(x).strip() if x is not None else "" for x in row]
                return i, headers
    return None


def _iter_sheet_records(ws, header_row: int, headers: List[str]) -> Iterable[Dict[str, Any]]:
    col_map = {idx: h for idx, h in enumerate(headers) if str(h or "").strip()}
    empty_streak = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            empty_streak += 1
            if empty_streak >= 50:
                break
            continue
        empty_streak = 0

        rec: Dict[str, Any] = {}
        for idx, v in enumerate(row):
            h = col_map.get(idx)
            if not h:
                continue
            rec[h] = v
        if all(v is None or str(v).strip() == "" for v in rec.values()):
            continue
        yield rec


def seed_printer_toner_from_workbook() -> None:
    # Best-effort: only attempt when repos/tables exist.
    # NOTE: We no longer bail out when data exists, because we may need to
    # backfill new columns (e.g. incidents.raw/rawHeaders).
    try:
        existing_incidents = DB.printer_toner_incidents.list()
        existing_entries = DB.printer_toner_entries.list()
        existing_exits = DB.printer_toner_exits.list()
        existing_min_qty = DB.printer_toner_min_qty.list()
    except Exception as e:
        logger.warning("Printer toner seed skipped (storage unavailable): %s", e)
        return

    repo_root = Path(__file__).resolve().parents[2]
    xlsm_path = repo_root / "public" / "suivie-incidents-imprimantes.xlsm"
    if not xlsm_path.exists():
        logger.info("Printer toner seed skipped (workbook missing): %s", xlsm_path)
        return

    wb = None
    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsm_path, data_only=True, keep_vba=True)

        # Incidents
        try:
            ws = wb["Incidents"]
            header = _find_header_row(ws, "Printer Name")
            if header is not None:
                header_row, headers = header
                raw_headers = [str(h).strip() for h in headers if str(h or "").strip()]

                for rec in _iter_sheet_records(ws, header_row, headers):
                    raw: Dict[str, Any] = {}
                    for k, v in rec.items():
                        txt = _to_text(v)
                        raw[str(k).strip()] = None if txt == "" else txt

                    site = _to_text(_first_present(rec, "Site"))
                    printer_name = _to_text(_first_present(rec, "Printer Name"))
                    demand_type = _to_text(_first_present(rec, "Type of demand"))
                    ticket = _to_text(_first_present(rec, "N° Ticket CBI", "N° Ticket", "Ticket"))
                    nature = _to_text(_first_present(rec, "Nature du Probleme", "Nature du Problème"))
                    serial = _to_text(_first_present(rec, "N° de Série Imprimante", "N° de Série"))
                    model = _to_text(_first_present(rec, "Model Imprimante", "Modèle Imprimante"))
                    claim_date = _to_date_iso(_first_present(rec, "Date Reclamation", "Date Réclamation"))
                    interv_date = _to_date_iso(_first_present(rec, "Date D'intervention CBI", "Date d'intervention CBI"))
                    duration = _to_text(_first_present(rec, "Duree de traitement ticket", "Duree de traitement ticket "))

                    status = "INTERVENUE" if interv_date else "NON_INTERVENUE"

                    if not (site or printer_name or ticket or nature):
                        continue

                    item_id = _gen_id("pti", site, printer_name, ticket, claim_date, nature)
                    item = models.PrinterTonerIncident(
                        id=item_id,
                        site=site or None,
                        printerName=printer_name or None,
                        demandType=demand_type or None,
                        ticketNumber=ticket or None,
                        problemNature=nature or None,
                        printerSerial=serial or None,
                        printerModel=model or None,
                        claimDate=claim_date,
                        interventionDate=interv_date,
                        duration=duration or None,
                        status=status,
                        raw=raw or None,
                        rawHeaders=raw_headers or None,
                    )
                    try:
                        if len(existing_incidents) == 0:
                            DB.printer_toner_incidents.create(item.id, lambda _id, it=item: it)
                        else:
                            def updater(current: models.PrinterTonerIncident, it=item):
                                d = current.model_dump()
                                d.update(it.model_dump(exclude_none=True))
                                return models.PrinterTonerIncident(**d)

                            try:
                                DB.printer_toner_incidents.update(item.id, updater)
                            except KeyError:
                                DB.printer_toner_incidents.create(item.id, lambda _id, it=item: it)
                    except ValueError:
                        continue
        except Exception as e:
            logger.warning("Printer toner incidents seed skipped: %s", e)

        # Entrées
        try:
            if len(existing_entries) == 0:
                ws = wb["Entrées"]
                header = _find_header_row(ws, "Date d'entrée")
                if header is not None:
                    header_row, headers = header
                    for rec in _iter_sheet_records(ws, header_row, headers):
                        d = _to_date_iso(_first_present(rec, "Date d'entrée", "Date d entree"))
                        article = _to_text(_first_present(rec, "Article"))
                        code = _to_text(_first_present(rec, "Code Artice", "Code Article"))
                        qty = _to_int(_first_present(rec, "Quantité", "Quantite"))
                        if not (article or qty):
                            continue
                        item_id = _gen_id("pte", d, article, code, qty)
                        item = models.PrinterTonerEntry(
                            id=item_id,
                            date=d,
                            article=article or None,
                            articleCode=code or None,
                            quantity=int(qty or 0),
                        )
                        try:
                            DB.printer_toner_entries.create(item.id, lambda _id, it=item: it)
                        except ValueError:
                            continue
        except Exception as e:
            logger.warning("Printer toner entries seed skipped: %s", e)

        # Sorties
        try:
            if len(existing_exits) == 0:
                ws = wb["Sorties"]
                header = _find_header_row(ws, "Date de sortie")
                if header is not None:
                    header_row, headers = header
                    for rec in _iter_sheet_records(ws, header_row, headers):
                        d = _to_date_iso(_first_present(rec, "Date de sortie"))
                        article = _to_text(_first_present(rec, "Nom Article", "Article"))
                        code = _to_text(_first_present(rec, "Code Artice", "Code Article"))
                        qty = _to_int(_first_present(rec, "Quantité", "Quantite"))
                        if not (article or qty):
                            continue
                        item_id = _gen_id("ptx", d, article, code, qty)
                        item = models.PrinterTonerExit(
                            id=item_id,
                            date=d,
                            article=article or None,
                            articleCode=code or None,
                            quantity=int(qty or 0),
                        )
                        try:
                            DB.printer_toner_exits.create(item.id, lambda _id, it=item: it)
                        except ValueError:
                            continue
        except Exception as e:
            logger.warning("Printer toner exits seed skipped: %s", e)

        # Min qty (AS3)
        try:
            if len(existing_min_qty) == 0:
                if "AS3" in wb.sheetnames:
                    ws = wb["AS3"]
                    header = _find_header_row(ws, "Référence") or _find_header_row(ws, "Reference")
                    if header is not None:
                        header_row, headers = header
                        current_ref = ""
                        for rec in _iter_sheet_records(ws, header_row, headers):
                            ref = _to_text(_first_present(rec, "Référence ", "Référence", "Reference"))
                            color = _to_text(_first_present(rec, "Couleur", "Color")).upper()
                            qty = _to_int(_first_present(rec, "Nombre de toner ", "Nombre de toner", "Nombre de toner  "))

                            if ref:
                                current_ref = ref
                            used_ref = ref or current_ref
                            if not (used_ref and color and qty):
                                continue

                            item_id = _gen_id("ptm", used_ref, color)
                            item = models.PrinterTonerMinQty(
                                id=item_id,
                                ref=used_ref,
                                color=color,
                                minQty=int(qty),
                            )
                            try:
                                DB.printer_toner_min_qty.create(item.id, lambda _id, it=item: it)
                            except ValueError:
                                continue
        except Exception as e:
            logger.warning("Printer toner min qty seed skipped: %s", e)
    except Exception as e:
        logger.warning("Printer toner seed skipped (cannot read workbook): %s", e)
        return
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def sync_asset_statuses_from_assignments() -> None:
    """Keep assets.status consistent with assignments.

    Assets referenced by an Active assignment become Assigned.
    Assets not referenced by any Active assignment become Available.

    Does not override InRepair/Retired.
    """

    try:
        assets = DB.assets.list()
        assignments = DB.assignments.list()
    except Exception as e:
        logger.warning("Sync skipped (storage unavailable): %s", e)
        return

    active_asset_ids = {
        str(a.assetId)
        for a in assignments
        if a.assetId and (a.status or "Active") == "Active"
    }

    changed = 0
    for asset in assets:
        aid = str(asset.id)
        cur = getattr(asset, "status", None)
        if cur in {"InRepair", "Retired"}:
            continue

        desired = "Assigned" if aid in active_asset_ids else "Available"
        if cur != desired:
            def updater(current: models.Asset):
                d = current.model_dump()
                d["status"] = desired
                return models.Asset(**d)

            try:
                DB.assets.update(aid, updater)
                changed += 1
            except Exception:
                continue

    logger.info(
        "Asset status sync complete: active_assets=%s changed=%s",
        len(active_asset_ids),
        changed,
    )


def seed_data() -> None:
    try:
        # If SQL Server is enabled but not reachable, don't crash app startup.
        _ = DB.departments.list()
    except Exception as e:
        logger.warning("Seed skipped (storage unavailable): %s", e)
        return

    # Cleanup legacy placeholder sites that represented IT codes only.
    # We now store real sites (AS1/AS2/...) with a separate `codeIt` field.
    for legacy_site_id in ("site-SEB", "site-BOK", "site-MA6", "site-MA7"):
        try:
            DB.sites.delete(legacy_site_id)
        except Exception:
            pass

    sites_seed = [
        models.Site(
            id="site-AS1",
            name="AS1",
            codeIt="SEB",
            location="Aïn Sebaâ",
            zone="Zone industrielle",
            city="Casablanca",
        ),
        models.Site(
            id="site-AS2",
            name="AS2",
            codeIt="SEB",
            location="Herbili",
            zone="Zone industrielle",
            city="Casablanca",
        ),
        models.Site(
            id="site-AS3",
            name="AS3",
            codeIt="AS3",
            location="Bouznika",
            zone="Zone industrielle",
            city="Bouznika",
        ),
        models.Site(
            id="site-ECO-PARK-A",
            name="Ecopark A",
            codeIt="MA7",
            location="Had Soualem",
            zone="Parc industriel",
            city="Had Soualem",
        ),
        models.Site(
            id="site-BOUSKOURA-1",
            name="Bouskoura 1",
            codeIt="BOK",
            location="Bouskoura",
            zone="Zone industrielle Bouskoura",
            city="Bouskoura",
        ),
        models.Site(
            id="site-BOUSKOURA-2",
            name="Bouskoura 2",
            codeIt="BOK",
            location="Bouskoura",
            zone="Zone industrielle Bouskoura",
            city="Bouskoura",
        ),
        models.Site(
            id="site-BOUSKOURA-3",
            name="Bouskoura 3",
            codeIt="BOK",
            location="Ouled Saleh",
            zone="Zone industrielle",
            city="Bouskoura",
        ),
        models.Site(
            id="site-SYSAPP",
            name="Sysapp",
            codeIt="MA7",
            location="Berrechid",
            zone="Zone industrielle (Hay Sinaâ)",
            city="Berrechid",
        ),
        models.Site(
            id="site-BERRECHID-2",
            name="Berrechid 2",
            codeIt="MA6",
            location="Berrechid",
            zone="Zone industrielle Berrechid",
            city="Berrechid",
        ),
        models.Site(
            id="site-AGADIR",
            name="Agadir",
            codeIt="MAG",
            location="Agadir",
            zone="Zone industrielle Agadir",
            city="Agadir",
        ),
    ]

    categories_seed = [
        models.Category(id="cat-MON", name="Monitor"),
        # User asset types
        models.Category(id="cat-WKS", name="Workstation"),
        models.Category(id="cat-NBK", name="Notebook"),
        models.Category(id="cat-PRI", name="Printer"),
        models.Category(id="cat-DOC", name="Docking Station"),
        models.Category(id="cat-APS", name="APs"),
        models.Category(id="cat-SCN", name="Scanner"),
        models.Category(id="cat-KAB", name="KABA"),
        models.Category(id="cat-CIS", name="Cisco"),
    ]

    departments_seed = [
        models.Department(
            id="dept-001",
            name="IT",
            code="IT",
            head="John Smith",
            members=12,
            description="Infrastructure & support",
        ),
        models.Department(
            id="dept-002",
            name="Finance",
            code="FIN",
            head="Sarah Johnson",
            members=8,
            description="Budgets & procurement",
        ),
        models.Department(
            id="dept-003",
            name="LOG",
            code="LOG",
            head="Karim Ben Ali",
            members=22,
            description="Warehousing & deliveries",
        ),
        models.Department(id="dept-004", name="PROD", code="PROD"),
        models.Department(id="dept-005", name="FRM", code="FRM"),
        models.Department(id="dept-006", name="ENG", code="ENG"),
        models.Department(id="dept-007", name="HR", code="HR"),
        models.Department(id="dept-008", name="MET", code="MET"),
        models.Department(id="dept-009", name="QL", code="QL"),
        models.Department(id="dept-010", name="FS", code="FS"),
        models.Department(id="dept-011", name="Maintenance", code="MAINT"),
        models.Department(id="dept-012", name="Laboratory", code="LAB"),
        models.Department(id="dept-013", name="Tooling", code="TOOL"),
        models.Department(id="dept-014", name="Log (zone principale)", code="LOG-ZP"),
        models.Department(id="dept-015", name="Log (zone2)", code="LOG-Z2"),
        models.Department(id="dept-016", name="Log bureau", code="LOG-BUR"),
        models.Department(id="dept-017", name="HR (transp)", code="HR-TR"),
    ]

    suppliers_seed = [
        models.Supplier(id="sup-001", name="TechSupplier Inc.", contact="sales@techsupplier.com"),
        models.Supplier(id="sup-002", name="OfficeGear", contact="contact@officegear.com"),
        models.Supplier(id="sup-003", name="NetworkPro", contact="support@networkpro.example"),
    ]

    demo_password = (os.environ.get("PFE_DEMO_PASSWORD") or "123456").strip() or "123456"
    demo_hash = hash_password(demo_password)

    demo_signature = (
        "data:image/svg+xml;utf8,"
        "<svg xmlns='http://www.w3.org/2000/svg' width='200' height='60'>"
        "<path d='M10,40C40,10,80,70,120,30S190,40,190,40' stroke='black' stroke-width='3' fill='none'/>"
        "</svg>"
    )

    users_seed = [
        models.UserDB(
            id="usr-001",
            name="Admin User",
            email="admin@leoni.example",
            role="Admin",
            avatarUrl=None,
            signatureNumber="100001",
            signatureData=demo_signature,
            passwordHash=demo_hash,
        ),
        models.UserDB(
            id="usr-002",
            name="Technician",
            email="tech@leoni.example",
            role="Technician",
            avatarUrl=None,
            signatureNumber="100002",
            signatureData=demo_signature,
            passwordHash=demo_hash,
        ),
        models.UserDB(
            id="usr-003",
            name="Manager",
            email="manager@leoni.example",
            role="Manager",
            avatarUrl=None,
            signatureNumber="100003",
            signatureData=demo_signature,
            passwordHash=demo_hash,
        ),
        models.UserDB(
            id="usr-004",
            name="Reader",
            email="reader@leoni.example",
            role="Reader",
            avatarUrl=None,
            signatureNumber="100004",
            signatureData=demo_signature,
            passwordHash=demo_hash,
        ),
        models.UserDB(
    id="usr-005",
    name="Maria Haddouch",
    email="mariaa.haddouch@gmail.com",
    role="Manager",  # ou "Technician", "Reader", etc. selon le besoin
    avatarUrl=None,
    signatureNumber="100005",
    signatureData=demo_signature,
    passwordHash=demo_hash,
),
    ]

    vendors_seed = [
        models.Vendor(
            id="ven-001",
            name="TechSupplier Inc.",
            category="IT Hardware",
            status="PREFERRED",
            email="sales@techsupplier.com",
            phone="+216 00 000 000",
            totalSpend=154000,
            activeContracts=2,
            rating=4.8,
            compliant=True,
        ),
        models.Vendor(
            id="ven-002",
            name="OfficeGear",
            category="Office/Peripherals",
            status="APPROVED",
            email="contact@officegear.com",
            phone="+216 00 111 111",
            totalSpend=62000,
            activeContracts=1,
            rating=4.4,
            compliant=True,
        ),
        models.Vendor(
            id="ven-003",
            name="NetworkPro",
            category="Networking",
            status="APPROVED",
            email="support@networkpro.example",
            phone="+216 00 222 222",
            totalSpend=98000,
            activeContracts=2,
            rating=4.6,
            compliant=True,
        ),
    ]

    # Always ensure master data exists.
    # For departments, use an upsert-like seed so updates are applied on re-run.
    for d in departments_seed:
        did = str(d.id)
        try:
            DB.departments.create(did, lambda _id, it=d: it)
        except ValueError:
            DB.departments.update(did, lambda _cur, it=d: it)
    for s in sites_seed:
        sid = str(s.id)
        try:
            DB.sites.create(sid, lambda _id, it=s: it)
        except ValueError:
            DB.sites.update(sid, lambda _cur, it=s: it)
    DB.categories.seed(categories_seed, get_id=lambda c: c.id)
    DB.suppliers.seed(suppliers_seed, get_id=lambda s: s.id)
    DB.users.seed(users_seed, get_id=lambda u: u.id)
    DB.vendors.seed(vendors_seed, get_id=lambda v: v.id)

    # Ensure demo accounts have a password hash even if they existed already.
    for u in users_seed:
        try:
            current = DB.users.get(u.id)
            if current is None:
                continue
            cur_hash = str(getattr(current, "passwordHash", "") or "")
            if cur_hash:
                continue

            def updater(existing: models.UserDB):
                d = existing.model_dump()
                d["passwordHash"] = demo_hash
                return models.UserDB.model_validate(d)

            DB.users.update(u.id, updater)
        except Exception:
            continue

    # Ensure demo accounts have a signatureData even if they existed already.
    for u in users_seed:
        try:
            current = DB.users.get(u.id)
            if current is None:
                continue
            cur_sig = str(getattr(current, "signatureData", "") or "").strip()
            if cur_sig:
                continue

            def updater(existing: models.UserDB):
                d = existing.model_dump()
                d["signatureData"] = demo_signature
                return models.UserDB.model_validate(d)

            DB.users.update(u.id, updater)
        except Exception:
            continue

    # Seed a richer asset catalog only when the DB is (almost) empty.
    try:
        existing_assets = DB.assets.list()
        should_seed_assets = len(existing_assets) < 5
    except Exception:
        should_seed_assets = True

    if should_seed_assets:
        assets_seed = [
            # Notebooks
            models.Asset(
                id="asset-001",
                assetTag="NB-MA6-0001",
                serialNumber="SN-NB-MA6-0001",
                macAddress="00:11:22:33:44:01",
                model="Dell Latitude 5420",
                type=None,
                deviceProfile=None,
                category="Notebook",
                supplier="TechSupplier Inc.",
                site="Berrechid 2",
                status="Available",
                warrantyEndDate="2026-12-31",
                acquisitionDate="2024-01-15",
                value=1200,
            ),
            models.Asset(
                id="asset-002",
                assetTag="NB-MA7-0002",
                serialNumber="SN-NB-MA7-0002",
                macAddress="00:11:22:33:44:02",
                model="HP ProBook 450 G8",
                type=None,
                deviceProfile=None,
                category="Notebook",
                supplier="OfficeGear",
                site="Ecopark A",
                status="Assigned",
                warrantyEndDate="2027-06-30",
                acquisitionDate="2025-06-15",
                value=980,
            ),
            # Workstations
            models.Asset(
                id="asset-003",
                assetTag="WS-SEB-0003",
                serialNumber="SN-WS-SEB-0003",
                macAddress=None,
                model="Lenovo ThinkStation P360",
                type=None,
                deviceProfile=None,
                category="Workstation",
                supplier="TechSupplier Inc.",
                site="AS1",
                status="Available",
                warrantyEndDate="2027-03-31",
                acquisitionDate="2025-03-20",
                value=1600,
            ),
            models.Asset(
                id="asset-004",
                assetTag="WS-BOK-0004",
                serialNumber="SN-WS-BOK-0004",
                macAddress=None,
                model="Dell OptiPlex 7010",
                type=None,
                deviceProfile=None,
                category="Workstation",
                supplier="OfficeGear",
                site="Bouskoura 1",
                status="InRepair",
                warrantyEndDate="2026-10-01",
                acquisitionDate="2024-10-01",
                value=850,
            ),
            # Monitors
            models.Asset(
                id="asset-005",
                assetTag="MON-MA6-0005",
                serialNumber="SN-MON-MA6-0005",
                macAddress=None,
                model='Dell P2422H 24"',
                type=None,
                deviceProfile=None,
                category="Monitor",
                supplier="OfficeGear",
                site="Berrechid 2",
                status="Available",
                warrantyEndDate="2027-12-31",
                acquisitionDate="2025-12-01",
                value=190,
            ),
            models.Asset(
                id="asset-006",
                assetTag="MON-AS3-0006",
                serialNumber="SN-MON-AS3-0006",
                macAddress=None,
                model='HP E24 G5 24"',
                type=None,
                deviceProfile=None,
                category="Monitor",
                supplier="OfficeGear",
                site="AS3",
                status="Assigned",
                warrantyEndDate="2028-01-15",
                acquisitionDate="2026-01-15",
                value=175,
            ),
            # Docking Stations
            models.Asset(
                id="asset-007",
                assetTag="DOC-MA6-0007",
                serialNumber="SN-DOC-MA6-0007",
                macAddress=None,
                model="HP USB-C Dock G5",
                type=None,
                deviceProfile=None,
                category="Docking Station",
                supplier="OfficeGear",
                site="Berrechid 2",
                status="Available",
                warrantyEndDate="2027-07-01",
                acquisitionDate="2025-07-01",
                value=140,
            ),
            models.Asset(
                id="asset-008",
                assetTag="DOC-MA7-0008",
                serialNumber="SN-DOC-MA7-0008",
                macAddress=None,
                model="Dell WD19",
                type=None,
                deviceProfile=None,
                category="Docking Station",
                supplier="TechSupplier Inc.",
                site="Ecopark A",
                status="Available",
                warrantyEndDate="2027-02-28",
                acquisitionDate="2025-03-01",
                value=165,
            ),
            # Printers
            models.Asset(
                id="asset-009",
                assetTag="PRI-SEB-0009",
                serialNumber="SN-PRI-SEB-0009",
                macAddress=None,
                model="HP LaserJet Pro M404dn",
                type=None,
                deviceProfile=None,
                category="Printer",
                supplier="OfficeGear",
                site="AS1",
                status="Available",
                warrantyEndDate="2026-11-30",
                acquisitionDate="2024-12-01",
                value=320,
            ),
            models.Asset(
                id="asset-010",
                assetTag="PRI-BOK-0010",
                serialNumber="SN-PRI-BOK-0010",
                macAddress=None,
                model="Brother HL-L6400DW",
                type=None,
                deviceProfile=None,
                category="Printer",
                supplier="OfficeGear",
                site="Bouskoura 1",
                status="Retired",
                warrantyEndDate="2025-12-31",
                acquisitionDate="2023-01-10",
                value=0,
            ),
            # APs
            models.Asset(
                id="asset-011",
                assetTag="AP-MA6-0011",
                serialNumber="SN-AP-MA6-0011",
                macAddress="10:20:30:40:50:11",
                model="Cisco Aironet 1832i",
                type=None,
                deviceProfile=None,
                category="APs",
                supplier="NetworkPro",
                site="Berrechid 2",
                status="Available",
                warrantyEndDate="2028-05-31",
                acquisitionDate="2026-02-01",
                value=420,
            ),
            models.Asset(
                id="asset-012",
                assetTag="AP-MA7-0012",
                serialNumber="SN-AP-MA7-0012",
                macAddress="10:20:30:40:50:12",
                model="Cisco Catalyst 9105AXI",
                type=None,
                deviceProfile=None,
                category="APs",
                supplier="NetworkPro",
                site="Ecopark A",
                status="Assigned",
                warrantyEndDate="2029-01-31",
                acquisitionDate="2026-03-01",
                value=530,
            ),
            # Scanners
            models.Asset(
                id="asset-013",
                assetTag="SCN-SEB-0013",
                serialNumber="SN-SCN-SEB-0013",
                macAddress=None,
                model="Fujitsu ScanSnap iX1600",
                type=None,
                deviceProfile=None,
                category="Scanner",
                supplier="OfficeGear",
                site="AS1",
                status="Available",
                warrantyEndDate="2027-09-30",
                acquisitionDate="2025-10-01",
                value=380,
            ),
            # KABA
            models.Asset(
                id="asset-014",
                assetTag="KAB-AS3-0014",
                serialNumber="SN-KAB-AS3-0014",
                macAddress=None,
                model="KABA Access Controller",
                type=None,
                deviceProfile=None,
                category="KABA",
                supplier="TechSupplier Inc.",
                site="AS3",
                status="Available",
                warrantyEndDate="2028-12-31",
                acquisitionDate="2026-01-01",
                value=900,
            ),
            # Cisco (generic)
            models.Asset(
                id="asset-015",
                assetTag="CIS-MA6-0015",
                serialNumber="SN-CIS-MA6-0015",
                macAddress=None,
                model="Cisco Switch Catalyst 9200L",
                type=None,
                deviceProfile=None,
                category="Cisco",
                supplier="NetworkPro",
                site="Berrechid 2",
                status="Available",
                warrantyEndDate="2030-12-31",
                acquisitionDate="2026-02-10",
                value=2100,
            ),
        ]

        DB.assets.seed(assets_seed, get_id=lambda a: a.id)

    DB.audit_logs.seed(
        [
            models.AuditLog(
                id="log-001",
                timestamp="2026-01-01 09:00",
                user="Admin User",
                userRole="Admin",
                userInitials="AU",
                action="CREATE",
                entity="Department",
                entityId="dept-001",
                description="Created IT Department",
                result="Success",
                ip="127.0.0.1",
                details={"source": "seed"},
            )
        ],
        get_id=lambda l: l.id,
    )

    # Seed printer toner/consumables data from the provided workbook (if any).
    try:
        seed_printer_toner_from_workbook()
    except Exception as e:
        logger.warning("Printer toner seed failed: %s", e)
