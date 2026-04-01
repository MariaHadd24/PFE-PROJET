from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


def load_env_file(path: Path) -> None:
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        os.environ[key] = value


def q_ident(name: str) -> str:
    return f"[{name.replace(']', ']]')}]"


def q_table(schema: str, table: str) -> str:
    return f"{q_ident(schema)}.{q_ident(table)}"


def normalize(name: str) -> str:
    s = name.strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    s = re.sub(r"[\s\-/\.]+", "_", s)
    s = re.sub(r"[^a-z0-9_]+", "", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def to_db_value(value: Any) -> Any:
    if value is None:
        return None

    if isinstance(value, datetime):
        # keep datetime for DATETIME2 columns
        return value

    if isinstance(value, date):
        # keep date for DATE columns
        return value

    # openpyxl can return empty strings; treat as NULL
    if isinstance(value, str):
        v = value.strip()
        if v == "":
            return None
        # Try ISO date parsing for convenience
        try:
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
                return datetime.fromisoformat(v).date()
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(:\d{2})?", v):
                return datetime.fromisoformat(v.replace(" ", "T"))
        except Exception:
            pass
        # Normalize boolean-like strings
        vl = v.lower()
        if vl in ("true", "yes", "y", "1", "oui"):
            return True
        if vl in ("false", "no", "n", "0", "non"):
            return False
        return v

    return value


def as_date(value: Any) -> Optional[date]:
    v = to_db_value(value)
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        try:
            return datetime.fromisoformat(v).date()
        except Exception:
            return None
    return None


def first_present(rec: Dict[str, Any], *keys: str) -> Any:
    for k in keys:
        if k in rec and rec[k] is not None and str(rec[k]).strip() != "":
            return rec[k]
    norm_rec = {normalize(k): v for k, v in rec.items()}
    for k in keys:
        nk = normalize(k)
        if nk in norm_rec and norm_rec[nk] is not None and str(norm_rec[nk]).strip() != "":
            return norm_rec[nk]
    return None


def gen_id(prefix: str, *parts: Any, max_len: int = 64) -> str:
    raw = "_".join(str(p) for p in parts if p is not None and str(p).strip() != "")
    raw = normalize(raw)
    if not raw:
        raw = "x"
    out = f"{prefix}-{raw}"
    if len(out) <= max_len:
        return out
    # Keep deterministic suffix
    return out[: max_len - 9] + "-" + out[-8:]


@dataclass(frozen=True)
class ColumnInfo:
    name: str
    is_nullable: bool


def get_table_columns(cur, schema: str) -> Dict[str, List[ColumnInfo]]:
    cur.execute(
        """
        SELECT TABLE_NAME, COLUMN_NAME, IS_NULLABLE
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = ?
        ORDER BY TABLE_NAME, ORDINAL_POSITION
        """,
        (schema,),
    )
    out: Dict[str, List[ColumnInfo]] = {}
    for table_name, column_name, is_nullable in cur.fetchall():
        out.setdefault(table_name, []).append(
            ColumnInfo(name=column_name, is_nullable=(str(is_nullable).upper() == "YES"))
        )
    return out


def iter_sheet_records(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
    # Find a header row in first 30 rows
    header_row_idx: Optional[int] = None
    headers: List[str] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        values = [v for v in row if v is not None and str(v).strip() != ""]
        if len(values) >= 2:
            header_row_idx = i
            headers = [str(v).strip() if v is not None else "" for v in row]
            break

    if header_row_idx is None:
        return [], []

    # Build map of column index -> header
    col_map: Dict[int, str] = {}
    for idx, h in enumerate(headers):
        h = (h or "").strip()
        if not h:
            continue
        col_map[idx] = h

    records: List[Dict[str, Any]] = []
    empty_streak = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Stop if we see many empty rows in a row
        if all(v is None or str(v).strip() == "" for v in row):
            empty_streak += 1
            if empty_streak >= 50:
                break
            continue
        empty_streak = 0

        rec: Dict[str, Any] = {}
        for idx, v in enumerate(row):
            if idx not in col_map:
                continue
            rec[col_map[idx]] = to_db_value(v)

        # Skip records with no meaningful data
        if all(v is None for v in rec.values()):
            continue

        records.append(rec)

    return list(col_map.values()), records


def guess_table(sheet_name: str, headers: List[str], tables: Dict[str, List[ColumnInfo]]) -> Optional[str]:
    sheet_norm = normalize(sheet_name)
    header_norms = {normalize(h) for h in headers if h}

    # Direct name match
    for t in tables.keys():
        if normalize(t) == sheet_norm:
            return t

    # Score by overlapping columns
    best: Tuple[int, str] = (0, "")
    for t, cols in tables.items():
        col_norms = {normalize(c.name) for c in cols}
        score = len(header_norms & col_norms)
        if score > best[0]:
            best = (score, t)

    if best[0] >= 2:
        return best[1]

    return None


def build_column_mapping(headers: List[str], table_cols: List[ColumnInfo]) -> Dict[str, str]:
    col_by_norm = {normalize(c.name): c.name for c in table_cols}
    mapping: Dict[str, str] = {}
    for h in headers:
        hn = normalize(h)
        if hn in col_by_norm:
            mapping[h] = col_by_norm[hn]
    return mapping


def required_columns(table_cols: List[ColumnInfo]) -> List[str]:
    # id is always required, but we treat it separately
    return [c.name for c in table_cols if (not c.is_nullable) and c.name != "id"]


def row_missing_required(row: Dict[str, Any], required: List[str]) -> List[str]:
    missing: List[str] = []
    for col in required:
        if col not in row or row[col] is None:
            missing.append(col)
    return missing


def upsert_by_id(cur, schema: str, table: str, row: Dict[str, Any]) -> str:
    id_value = row.get("id")
    if id_value is None or str(id_value).strip() == "":
        raise ValueError("missing id")

    columns = [c for c in row.keys() if c != "id"]

    cur.execute(
        f"SELECT 1 FROM {q_table(schema, table)} WHERE {q_ident('id')} = ?",
        (id_value,),
    )
    exists = cur.fetchone() is not None

    if exists:
        if not columns:
            return "skip"
        set_sql = ", ".join(f"{q_ident(c)} = ?" for c in columns)
        params = [row[c] for c in columns] + [id_value]
        cur.execute(
            f"UPDATE {q_table(schema, table)} SET {set_sql} WHERE {q_ident('id')} = ?",
            params,
        )
        return "update"

    insert_cols = ["id"] + columns
    col_sql = ", ".join(q_ident(c) for c in insert_cols)
    ph_sql = ", ".join(["?"] * len(insert_cols))
    params = [id_value] + [row[c] for c in columns]
    cur.execute(
        f"INSERT INTO {q_table(schema, table)} ({col_sql}) VALUES ({ph_sql})",
        params,
    )
    return "insert"


def ensure_category(cur, schema: str, name: str) -> None:
    n = str(name or "").strip()
    if not n:
        return
    row = {"id": gen_id("cat", n), "name": n}
    upsert_by_id(cur, schema=schema, table="categories", row=row)


def ensure_site(cur, schema: str, name: str) -> None:
    n = str(name or "").strip()
    if not n:
        return
    # location is required by schema; use name as a reasonable default
    row = {"id": gen_id("site", n), "name": n, "location": n}
    upsert_by_id(cur, schema=schema, table="sites", row=row)


def ensure_supplier(cur, schema: str, name: str) -> None:
    n = str(name or "").strip()
    if not n:
        return
    # contact is required by schema; use placeholder
    row = {"id": gen_id("sup", n), "name": n, "contact": "Unknown"}
    upsert_by_id(cur, schema=schema, table="suppliers", row=row)


def get_existing_asset_id_by_serial(cur, schema: str, serial_number: str) -> Optional[str]:
    cur.execute(
        f"SELECT {q_ident('id')} FROM {q_table(schema, 'assets')} WHERE {q_ident('serialNumber')} = ?",
        (serial_number,),
    )
    row = cur.fetchone()
    return str(row[0]) if row else None


def asset_tag_available(cur, schema: str, asset_tag: str, asset_id: Optional[str]) -> bool:
    if asset_id:
        cur.execute(
            f"SELECT 1 FROM {q_table(schema, 'assets')} WHERE {q_ident('assetTag')} = ? AND {q_ident('id')} <> ?",
            (asset_tag, asset_id),
        )
    else:
        cur.execute(
            f"SELECT 1 FROM {q_table(schema, 'assets')} WHERE {q_ident('assetTag')} = ?",
            (asset_tag,),
        )
    return cur.fetchone() is None


def upsert_asset(cur, schema: str, asset: Dict[str, Any]) -> Tuple[str, str]:
    # Returns (action, asset_id)
    serial_number = str(asset.get("serialNumber") or "").strip()
    if not serial_number:
        raise ValueError("missing serialNumber")

    existing_id = get_existing_asset_id_by_serial(cur, schema=schema, serial_number=serial_number)
    if existing_id:
        asset["id"] = existing_id

    if "id" not in asset or not asset["id"]:
        asset["id"] = gen_id("asset", serial_number)

    # Keep assetTag unique
    asset_tag = str(asset.get("assetTag") or "").strip()
    if not asset_tag:
        asset_tag = f"SN-{serial_number}"
        asset["assetTag"] = asset_tag

    if not asset_tag_available(cur, schema=schema, asset_tag=asset_tag, asset_id=existing_id):
        asset_tag = f"SN-{serial_number}"
        asset["assetTag"] = asset_tag

    action = upsert_by_id(cur, schema=schema, table="assets", row=asset)
    return action, str(asset["id"])


def import_inventory_ma6_sheet(
    cur,
    schema: str,
    sheet_name: str,
    headers: List[str],
    records: List[Dict[str, Any]],
    dry_run: bool,
) -> Dict[str, int]:
    stats = {"assets_insert": 0, "assets_update": 0, "assignments": 0, "rows": 0, "skipped": 0, "error": 0}

    sheet_norm = normalize(sheet_name)
    if sheet_norm not in ("ws_ma6", "nb", "print"):
        return stats

    for rec in records:
        stats["rows"] += 1

        if sheet_norm == "ws_ma6":
            serial = first_present(rec, "WS_SN")
            model = first_present(rec, "WS_model")
            mac = first_present(rec, "Mac Adress", "mac address", "mac")
            site = first_present(rec, "site", "Site")
            asset_tag = first_present(rec, "immo ws", "IMMO Number", "hostname")
            acquisition = as_date(first_present(rec, "Date d’acquisition", "Date d'acquisition", "Acquisition date"))
            assign_date = as_date(first_present(rec, "date d'affectation", "Date affectation", "Assignement date"))
            user_name = first_present(rec, "user", "Username")
            department = first_present(rec, "Service")
            eos = as_date(first_present(rec, "Date end of Support", "Date end of support", "EOS"))
            category = "Workstation"
        elif sheet_norm == "nb":
            serial = first_present(rec, "NB SN")
            model = first_present(rec, "Model NB")
            mac = first_present(rec, "Mac Adress")
            site = first_present(rec, "Site")
            asset_tag = first_present(rec, "IMMO Number", "Hostname")
            acquisition = as_date(first_present(rec, "Acquisition date", "Date d’acquisition", "Date d'acquisition"))
            assign_date = as_date(first_present(rec, "Assignement date", "Date affectation", "Date affectation"))
            user_name = first_present(rec, "Username")
            department = first_present(rec, "Service")
            eos = as_date(first_present(rec, "Date end of support", "Date end of Support", "EOS"))
            category = "Laptop"
        else:  # print
            serial = first_present(rec, "Printer SN", "printer sn", "SN")
            model = first_present(rec, "printer model", "Printer model", "Model")
            mac = first_present(rec, "IP")
            site = first_present(rec, "Site")
            asset_tag = first_present(rec, "printer name", "Printer name", "immo", "Immo Number")
            acquisition = as_date(first_present(rec, "date reception", "Date reception"))
            assign_date = as_date(first_present(rec, "Date affectation", "Date affectation"))
            user_name = None
            department = first_present(rec, "Area")
            eos = as_date(first_present(rec, "EOS"))
            category = "Printer"

        serial_s = str(serial).strip() if serial is not None else ""
        if not serial_s:
            stats["skipped"] += 1
            continue

        acquisition_d = acquisition or date.today()
        warranty_d = eos or acquisition_d

        status = "Assigned" if assign_date else "Available"

        asset: Dict[str, Any] = {
            "id": gen_id("asset", serial_s),
            "assetTag": str(asset_tag).strip() if asset_tag is not None else f"SN-{serial_s}",
            "serialNumber": serial_s,
            "macAddress": str(mac).strip() if mac is not None else None,
            "model": str(model).strip() if model is not None else "Unknown",
            "category": category,
            "supplier": "Unknown",
            "site": str(site).strip() if site is not None else "Unknown",
            "status": status,
            "warrantyEndDate": warranty_d.isoformat(),
            "acquisitionDate": acquisition_d.isoformat(),
            "value": 0.0,
        }

        try:
            if not dry_run:
                ensure_category(cur, schema=schema, name=asset["category"])
                ensure_site(cur, schema=schema, name=asset["site"])
                ensure_supplier(cur, schema=schema, name=asset["supplier"])

            if dry_run:
                asset_id = asset["id"]
                action = "insert"
            else:
                action, asset_id = upsert_asset(cur, schema=schema, asset=asset)

            if action == "insert":
                stats["assets_insert"] += 1
            elif action == "update":
                stats["assets_update"] += 1

            # Assignment
            if assign_date and user_name:
                asn = {
                    "id": gen_id("asn", asset_id, assign_date.isoformat()),
                    "assetId": asset_id,
                    "userName": str(user_name).strip(),
                    "department": str(department).strip() if department is not None else "",
                    "site": str(site).strip() if site is not None else "Unknown",
                    "startDate": assign_date.isoformat(),
                    "returnDate": None,
                    "status": "Active",
                    "approvedBy": None,
                }
                if not dry_run:
                    upsert_by_id(cur, schema=schema, table="assignments", row=asn)
                stats["assignments"] += 1
        except Exception:
            stats["error"] += 1

    return stats


def import_stock_inventory_sheet(
    cur,
    schema: str,
    sheet_name: str,
    headers: List[str],
    records: List[Dict[str, Any]],
    dry_run: bool,
) -> Dict[str, int]:
    stats = {"assets_insert": 0, "assets_update": 0, "movements": 0, "rows": 0, "skipped": 0, "error": 0}

    for rec in records:
        stats["rows"] += 1

        serial = first_present(rec, "SN")
        serial_s = str(serial).strip() if serial is not None else ""
        if not serial_s:
            stats["skipped"] += 1
            continue

        model = first_present(rec, "Model")
        mac = first_present(rec, "MAC")
        plant = first_present(rec, "Plant")
        asset_tag = first_present(rec, "Immo Number", "BCI", "VNC")
        comment = first_present(rec, "Comment", "comments")
        date_in = as_date(first_present(rec, "Date In"))
        date_out = as_date(first_present(rec, "Date Out"))
        pilote_in = first_present(rec, "Pilote")

        acquisition_d = date_in or date.today()
        warranty_d = acquisition_d

        status = "Available"
        if date_out:
            status = "Assigned"

        type_value = str(first_present(rec, "TYPE") or "").strip()
        category = (type_value or str(sheet_name)).strip() or sheet_name

        asset: Dict[str, Any] = {
            "id": gen_id("asset", serial_s),
            "assetTag": str(asset_tag).strip() if asset_tag is not None else f"SN-{serial_s}",
            "serialNumber": serial_s,
            "macAddress": str(mac).strip() if mac is not None else None,
            "model": str(model).strip() if model is not None else "Unknown",
            "type": type_value or None,
            "category": category,
            "supplier": "Unknown",
            "site": str(plant).strip() if plant is not None else "Unknown",
            "status": status,
            "warrantyEndDate": warranty_d.isoformat(),
            "acquisitionDate": acquisition_d.isoformat(),
            "value": 0.0,
        }

        # If the Excel doesn't provide TYPE (common for Cisco sheets), infer it from the model.
        if not asset.get("type"):
            try:
                from app.cisco import infer_cisco_network_device_type

                inferred = infer_cisco_network_device_type(
                    model=asset.get("model"),
                    description=comment,
                    category=asset.get("category"),
                    supplier=asset.get("supplier"),
                    current_type=asset.get("type"),
                )
                if inferred:
                    asset["type"] = inferred
            except Exception:
                # Best-effort: never break import if inference fails.
                pass

        try:
            if not dry_run:
                ensure_category(cur, schema=schema, name=asset["category"])
                ensure_site(cur, schema=schema, name=asset["site"])
                ensure_supplier(cur, schema=schema, name=asset["supplier"])

            if dry_run:
                asset_id = asset["id"]
                action = "insert"
            else:
                action, asset_id = upsert_asset(cur, schema=schema, asset=asset)

            if action == "insert":
                stats["assets_insert"] += 1
            elif action == "update":
                stats["assets_update"] += 1

            # Create stock movements from Date In / Date Out
            for movement_type, movement_date in (("Entry", date_in), ("Exit", date_out)):
                if not movement_date:
                    continue
                mov = {
                    "id": gen_id("mov", asset_id, movement_type, movement_date.isoformat()),
                    "assetId": asset_id,
                    "type": movement_type,
                    "sourceSite": None,
                    "destinationSite": str(plant).strip() if plant is not None else None,
                    "date": movement_date.isoformat(),
                    "user": str(pilote_in).strip() if pilote_in is not None else "",
                    "comment": str(comment).strip() if comment is not None else "",
                }
                if not dry_run:
                    upsert_by_id(cur, schema=schema, table="stock_movements", row=mov)
                stats["movements"] += 1
        except Exception:
            stats["error"] += 1

    return stats


def import_workbook(cur, schema: str, path: Path, tables: Dict[str, List[ColumnInfo]], dry_run: bool) -> Dict[str, Dict[str, int]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    stats: Dict[str, Dict[str, int]] = {}

    for ws in wb.worksheets:
        headers, records = iter_sheet_records(ws)
        if not headers or not records:
            continue

        # Special handling for the user's MA6 workbooks
        lower_path = path.name.lower()

        # Route the "Stock Inventory" workbook first (its name also contains "inventory").
        if "stock inventory" in lower_path or ("stock" in lower_path and "inventory" in lower_path):
            s = import_stock_inventory_sheet(cur, schema, ws.title, headers, records, dry_run=dry_run)
            if s["rows"] > 0:
                stats[f"stock:{ws.title}"] = s
            continue

        if "inventory" in lower_path:
            s = import_inventory_ma6_sheet(cur, schema, ws.title, headers, records, dry_run=dry_run)
            if s["rows"] > 0:
                stats[f"inventory:{ws.title}"] = s
            continue

        # Generic table import (only when columns match table columns)
        table = guess_table(ws.title, headers, tables)
        if not table:
            continue

        table_cols = tables[table]
        mapping = build_column_mapping(headers, table_cols)
        required = required_columns(table_cols)

        s = stats.setdefault(table, {"insert": 0, "update": 0, "skip": 0, "error": 0, "rows": 0})

        for rec in records:
            s["rows"] += 1
            row: Dict[str, Any] = {}
            for src, dest in mapping.items():
                row[dest] = rec.get(src)

            # Allow already-correct keys too (if Excel headers already match column names)
            for k, v in rec.items():
                if k in (c.name for c in table_cols):
                    row[k] = v

            # Normalize "ID" / "Id" column header
            if "id" not in row:
                for k in list(row.keys()):
                    if normalize(k) == "id":
                        row["id"] = row.pop(k)
                        break

            missing = row_missing_required(row, required)
            if missing:
                s["error"] += 1
                continue

            try:
                if dry_run:
                    action = "insert"
                else:
                    action = upsert_by_id(cur, schema=schema, table=table, row=row)
                s[action] = s.get(action, 0) + 1
            except Exception:
                s["error"] += 1

    return stats


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Excel workbooks into SQL Server (upsert by id).")
    parser.add_argument(
        "--dir",
        default=str(Path("database/import")),
        help="Directory containing .xlsx files (default: database/import)",
    )
    parser.add_argument(
        "--files",
        nargs="*",
        default=None,
        help="Optional explicit list of .xlsx files. If omitted, imports all .xlsx in --dir.",
    )
    parser.add_argument("--schema", default="dbo")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--env", default=str(Path("backend/.env.sqlserver")), help="Env file to load")

    args = parser.parse_args()

    env_path = Path(args.env)
    if env_path.exists():
        load_env_file(env_path)

    repo_root = Path(__file__).resolve().parents[2]
    backend_dir = repo_root / "backend"
    sys.path.insert(0, str(backend_dir))

    from app.sqlserver import connect

    if args.files:
        files = [Path(p) for p in args.files]
    else:
        import_dir = Path(args.dir)
        files = sorted(import_dir.glob("*.xlsx"))

    if not files:
        print("No .xlsx files found.")
        return 2

    cn = connect()
    cur = cn.cursor()

    tables = get_table_columns(cur, schema=args.schema)

    all_stats: Dict[str, Dict[str, int]] = {}

    for path in files:
        if not path.exists():
            print(f"Missing file: {path}")
            return 2
        stats = import_workbook(cur, schema=args.schema, path=path, tables=tables, dry_run=args.dry_run)
        for table, s in stats.items():
            agg = all_stats.setdefault(table, {})
            for k, v in s.items():
                agg[k] = int(agg.get(k, 0)) + int(v)

    if args.dry_run:
        cn.rollback()
    else:
        cn.commit()

    print(json.dumps(all_stats, indent=2, default=str))

    cur.close()
    cn.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
