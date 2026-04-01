from __future__ import annotations

import argparse
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import httpx
from openpyxl import load_workbook


def _norm_key(s: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in s.strip()).strip("_")


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v).strip()


def _norm_serial(v: Any) -> str:
    s = _to_str(v)
    if not s:
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", s).upper()


def _read_sheet(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(v).strip() if v is not None else "" for v in headers_row]
    col_map = {i: h for i, h in enumerate(headers) if h}

    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rec: Dict[str, Any] = {}
        for i, v in enumerate(row):
            if i in col_map:
                rec[col_map[i]] = v
        rows.append(rec)

    return headers, rows


def _pick_first(rec: Dict[str, Any], keys: Iterable[str]) -> Optional[Any]:
    for k in keys:
        if k in rec and _to_str(rec.get(k)):
            return rec.get(k)
    return None


def main() -> int:
    ap = argparse.ArgumentParser(description="Import assignments from Inventory Excel into SQL Server via API")
    ap.add_argument(
        "--file",
        default="database/import/Inventory-MA6.xlsx",
        help="Path to Excel file (default: database/import/Inventory-MA6.xlsx)",
    )
    ap.add_argument(
        "--base-url",
        default="http://127.0.0.1:8001",
        help="Backend base URL (default: http://127.0.0.1:8001)",
    )
    ap.add_argument(
        "--replace",
        action="store_true",
        help="Delete all assignments before importing",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and report, but do not call the API",
    )
    args = ap.parse_args()

    xlsx = Path(args.file)
    if not xlsx.exists():
        raise SystemExit(f"Missing file: {xlsx}")

    base_url = args.base_url.rstrip("/")

    with httpx.Client(timeout=30.0) as client:
        # Sanity check
        if not args.dry_run:
            health = client.get(f"{base_url}/health")
            health.raise_for_status()
            j = health.json()
            print(f"health: status={j.get('status')} storage={j.get('storage')} sqlserver={j.get('sqlserver')}")

        # Load assets once for lookup
        assets_by_serial: Dict[str, str] = {}
        if not args.dry_run:
            assets_res = client.get(f"{base_url}/assets", params={"limit": 10000})
            assets_res.raise_for_status()
            assets = assets_res.json()
            for a in assets:
                serial = _norm_serial(a.get("serialNumber"))
                if serial:
                    assets_by_serial[serial] = str(a.get("id"))
            print(f"assets: loaded={len(assets)} mapped_by_serial={len(assets_by_serial)}")

        # Replace mode
        if args.replace and not args.dry_run:
            del_res = client.delete(f"{base_url}/assignments")
            del_res.raise_for_status()
            print(f"delete_all_assignments: {del_res.json()}")

        wb = load_workbook(xlsx, data_only=True)

        # Known target sheets and primary columns (based on prior inspection scripts)
        targets = {
            "ws_ma6": {
                "label": "WS MA6",
                "device_category": "Workstation",
                "user": ["user", "User", "FULL_NAME", "Full Name", "full_name"],
                "username": ["username", "Username"],
                "assign": [
                    "date d'affectation",
                    "Date affectation",
                    "Assignement date",
                    "Assignment date",
                    "assignment_date",
                ],
                "serial": ["WS_SN", "WS SN", "ws_sn", "Serial", "Serial Number"],
            },
            "nb": {
                "label": "NB",
                "device_category": "Notebook",
                "user": ["user", "User", "FULL_NAME", "Full Name", "full_name"],
                "username": ["username", "Username"],
                "assign": [
                    "date d'affectation",
                    "Date affectation",
                    "Assignement date",
                    "Assignment date",
                    "assignment_date",
                ],
                "serial": ["NB SN", "NB_SN", "nb_sn", "Serial", "Serial Number"],
            },
        }

        # Heuristic header->AssignmentCreate key mapping
        header_map = {
            # core
            "device_category": "device_category",
            "hostname": "hostname",
            "department": "department",
            "site": "site",
            "status": "status",
            "approvedby": "approvedBy",
            "approved_by": "approvedBy",
            "startdate": "startDate",
            "start_date": "startDate",
            "returndate": "returnDate",
            "return_date": "returnDate",
            "assignment_date": "assignment_date",
            "acquisition_date": "acquisition_date",
            "end_of_support_date": "end_of_support_date",
            # workstation/notebook fields
            "usb_status": "usb_status",
            "usb": "usb",
            "user": "user",
            "username": "username",
            "full_name": "full_name",
            "service": "service",
            "ws_sn": "ws_sn",
            "ws_model": "ws_model",
            "nb_sn": "nb_sn",
            "model_nb": "model_nb",
            "mac_address": "mac_address",
            "os": "os",
            "immo_ws": "immo_ws",
            "immo_number": "immo_number",
            "bci_ws": "bci_ws",
            "bci": "bci",
            "monitor_model": "monitor_model",
            "monitor_sn": "monitor_sn",
            "monitor_immo": "monitor_immo",
            "monitor_bci": "monitor_bci",
        }

        def map_record(rec: Dict[str, Any]) -> Dict[str, Any]:
            payload: Dict[str, Any] = {}
            for k, v in rec.items():
                nk = _norm_key(str(k))
                if nk in header_map:
                    out_k = header_map[nk]
                    sv = _to_str(v)
                    if sv:
                        payload[out_k] = sv
            return payload

        created = 0
        skipped = 0
        missing_asset = 0
        retried_without_asset = 0
        errors = 0

        for sheet_norm, cfg in targets.items():
            ws = None
            for sname in wb.sheetnames:
                if _norm_key(sname) == sheet_norm:
                    ws = wb[sname]
                    break
            if ws is None:
                print(f"sheet: missing {cfg['label']} ({sheet_norm})")
                continue

            _, rows = _read_sheet(ws)
            print(f"sheet: {cfg['label']} rows={len(rows)}")

            for rec in rows:
                assign_date = _pick_first(rec, cfg["assign"])  # type: ignore[index]
                user_val = _pick_first(rec, cfg["user"])  # type: ignore[index]
                username_val = _pick_first(rec, cfg["username"])  # type: ignore[index]

                if assign_date is None and user_val is None and username_val is None:
                    skipped += 1
                    continue

                payload = map_record(rec)

                # Force category per sheet unless the record already has it
                payload.setdefault("device_category", cfg["device_category"])  # type: ignore[index]

                # Fill key identity fields
                if user_val is not None and "userName" not in payload and "full_name" not in payload and "user" not in payload:
                    payload["userName"] = _to_str(user_val)
                if username_val is not None and "username" not in payload:
                    payload["username"] = _to_str(username_val)

                # Dates: prefer assignment_date
                if assign_date is not None:
                    payload.setdefault("assignment_date", _to_str(assign_date))
                    payload.setdefault("startDate", _to_str(assign_date))

                # Try resolve assetId by serial
                serial_val = _pick_first(rec, cfg["serial"])  # type: ignore[index]
                serial_norm = _norm_serial(serial_val)

                if serial_norm and assets_by_serial:
                    asset_id = assets_by_serial.get(serial_norm)
                    if asset_id:
                        payload["assetId"] = asset_id
                    else:
                        missing_asset += 1
                elif serial_norm:
                    # dry-run or no assets loaded
                    missing_asset += 1

                # Last cleanup: strip whitespace
                payload = {k: (v.strip() if isinstance(v, str) else v) for k, v in payload.items()}

                if args.dry_run:
                    created += 1
                    continue

                try:
                    res = client.post(f"{base_url}/assignments", json=payload)
                    if res.status_code >= 400:
                        # If asset is not available, retry without assetId so we still import the row.
                        if res.status_code == 409 and "Asset is not available" in (res.text or "") and payload.get("assetId"):
                            retry_payload = dict(payload)
                            retry_payload.pop("assetId", None)
                            res2 = client.post(f"{base_url}/assignments", json=retry_payload)
                            if res2.status_code < 400:
                                retried_without_asset += 1
                                created += 1
                                continue

                        errors += 1
                        print("ERROR", res.status_code, res.text[:400])
                        continue
                    created += 1
                except Exception as e:  # noqa: BLE001
                    errors += 1
                    print("ERROR", repr(e))

        print(
            "summary:",
            {
                "created": created,
                "skipped": skipped,
                "missing_asset_link": missing_asset,
                "retried_without_asset": retried_without_asset,
                "errors": errors,
            },
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
