from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, List, Optional, Tuple


def find_header_row(ws, max_rows: int = 30) -> Tuple[Optional[int], List[str]]:
    header_row_idx: Optional[int] = None
    headers: List[str] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        values = [v for v in row if v is not None and str(v).strip() != ""]
        if len(values) >= 2:
            header_row_idx = i
            headers = [str(v).strip() if v is not None else "" for v in row]
            break
    return header_row_idx, headers


def main() -> int:
    parser = argparse.ArgumentParser(description="Inspect Excel sheets + headers")
    parser.add_argument("files", nargs="*", default=None, help="xlsx files to inspect")
    parser.add_argument("--dir", default="database/import", help="Directory to scan if no files provided")
    args = parser.parse_args()

    if args.files:
        files = [Path(p) for p in args.files]
    else:
        files = sorted(Path(args.dir).glob("*.xlsx"))

    if not files:
        print("No xlsx files")
        return 2

    from openpyxl import load_workbook

    for path in files:
        print(f"\n=== {path} ===")
        wb = load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            header_idx, headers = find_header_row(ws)
            headers_clean = [h for h in headers if h]
            print(f"- sheet: {ws.title}")
            print(f"  header_row: {header_idx}")
            print(f"  headers({len(headers_clean)}): {headers_clean[:40]}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
