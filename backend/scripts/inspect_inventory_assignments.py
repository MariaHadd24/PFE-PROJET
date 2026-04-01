from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


def to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def main() -> int:
    path = Path("database/import/Inventory-MA6.xlsx")
    if not path.exists():
        print(f"Missing: {path}")
        return 2

    wb = load_workbook(path, data_only=True)

    def header_and_rows(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
        headers = [str(v).strip() if v is not None else "" for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        col_map = {i: h for i, h in enumerate(headers) if h}
        rows: List[Dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            rec: Dict[str, Any] = {}
            for i, v in enumerate(row):
                if i in col_map:
                    rec[col_map[i]] = v
            rows.append(rec)
        return headers, rows

    targets = {
        "ws_ma6": {
            "label": "WS MA6",
            "user": ["user"],
            "assign": ["date d'affectation"],
            "serial": ["WS_SN"],
        },
        "nb": {
            "label": "NB",
            "user": ["Username"],
            "assign": ["Assignement date", "Assignment date", "Date affectation", "date d'affectation"],
            "serial": ["NB SN"],
        },
    }

    def norm(s: str) -> str:
        return "".join(ch.lower() if ch.isalnum() else "_" for ch in s.strip()).strip("_")

    for sheet_norm, cols in targets.items():
        ws = None
        for sname in wb.sheetnames:
            if norm(sname) == sheet_norm:
                ws = wb[sname]
                break
        if ws is None:
            print(f"Missing sheet matching: {cols['label']}")
            continue
        _, rows = header_and_rows(ws)
        user_count = 0
        date_count = 0
        both = 0
        examples = 0
        for rec in rows:
            u = next((rec.get(k) for k in cols["user"] if to_str(rec.get(k))), None)
            d = next((rec.get(k) for k in cols["assign"] if rec.get(k) is not None and to_str(rec.get(k))), None)
            if u is not None:
                user_count += 1
            if d is not None:
                date_count += 1
            if u is not None and d is not None:
                both += 1
            if examples < 3 and (u is not None or d is not None):
                s = next((rec.get(k) for k in cols["serial"] if to_str(rec.get(k))), None)
                print(f"{cols['label']} example: serial={to_str(s)} user={to_str(u)} assign={to_str(d)}")
                examples += 1

        print(f"{cols['label']}: rows={len(rows)} user_nonempty={user_count} assign_nonempty={date_count} both={both}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
