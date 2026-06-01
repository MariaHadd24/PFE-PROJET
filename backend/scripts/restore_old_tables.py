from __future__ import annotations

import os
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def normalize(name: str) -> str:
    s = str(name or "").strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    s = re.sub(r"[\s\-/\.]+", "_", s)
    s = re.sub(r"[^a-z0-9_]+", "", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def to_int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    s = to_text(value).replace(",", ".")
    try:
        return int(round(float(s)))
    except Exception:
        return 0


def to_date_iso(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    s = to_text(value)
    if not s:
        return None

    try:
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
            return date.fromisoformat(s).isoformat()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}(:\d{2})?", s):
            return datetime.fromisoformat(s.replace(" ", "T")).date().isoformat()
    except Exception:
        pass

    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        dd, mm, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(yyyy, mm, dd).isoformat()
        except ValueError:
            return None

    return None


def first_present(rec: Dict[str, Any], *keys: str) -> Any:
    for k in keys:
        if k in rec and to_text(rec.get(k)):
            return rec.get(k)
    norm = {normalize(k): v for k, v in rec.items()}
    for k in keys:
        nk = normalize(k)
        if nk in norm and to_text(norm.get(nk)):
            return norm.get(nk)
    return None


def gen_id(prefix: str, *parts: Any, max_len: int = 64) -> str:
    raw = "_".join(to_text(p) for p in parts if to_text(p))
    raw = normalize(raw) or "x"
    out = f"{prefix}-{raw}"
    if len(out) <= max_len:
        return out
    return out[: max_len - 9] + "-" + out[-8:]


def find_header_row(ws, needle: str, *, max_rows: int = 60) -> Optional[Tuple[int, List[str]]]:
    n = needle.lower().strip()
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        for v in row:
            if v is None:
                continue
            if n in str(v).lower():
                return i, [str(x).strip() if x is not None else "" for x in row]
    return None


def iter_sheet_records(ws, header_row: int, headers: List[str]) -> Iterable[Dict[str, Any]]:
    col_map = {idx: h for idx, h in enumerate(headers) if str(h or "").strip()}
    empty_streak = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            empty_streak += 1
            if empty_streak >= 50:
                break
            continue
        empty_streak = 0

        rec: Dict[str, Any] = {}
        for idx, v in enumerate(row):
            header = col_map.get(idx)
            if header:
                rec[header] = v
        if rec:
            yield rec


def clear_repo(repo) -> int:
    removed = 0
    for item in list(repo.list()):
        try:
            repo.delete(str(item.id))
            removed += 1
        except Exception:
            continue
    return removed


def restore_printer_toner_workbook(repo_root: Path) -> dict[str, int]:
    from openpyxl import load_workbook

    from app import models
    from app.storage import DB

    workbook_path = repo_root / "database" / "import" / "suivie incidents imprimantes.xlsm"
    if not workbook_path.exists():
        workbook_path = repo_root / "public" / "suivie-incidents-imprimantes.xlsm"
    if not workbook_path.exists():
        print(f"Missing incidents workbook: {workbook_path}")
        return {"incidents": 0, "entries": 0, "exits": 0, "min_qty": 0}

    cleared = {
        "incidents": clear_repo(DB.printer_toner_incidents),
        "entries": clear_repo(DB.printer_toner_entries),
        "exits": clear_repo(DB.printer_toner_exits),
        "min_qty": clear_repo(DB.printer_toner_min_qty),
    }

    wb = load_workbook(workbook_path, data_only=True, keep_vba=True)

    # Incidents
    try:
        ws = wb["Incidents"]
        header = find_header_row(ws, "Printer Name")
        if header is not None:
            header_row, headers = header
            for rec in iter_sheet_records(ws, header_row, headers):
                raw: Dict[str, Any] = {}
                for k, v in rec.items():
                    txt = to_text(v)
                    raw[str(k).strip()] = None if txt == "" else txt

                site = to_text(first_present(rec, "Site"))
                printer_name = to_text(first_present(rec, "Printer Name"))
                demand_type = to_text(first_present(rec, "Type of demand"))
                ticket = to_text(first_present(rec, "N° Ticket CBI", "N° Ticket", "Ticket"))
                nature = to_text(first_present(rec, "Nature du Probleme", "Nature du Problème"))
                serial = to_text(first_present(rec, "N° de Série Imprimante", "N° de Série"))
                model = to_text(first_present(rec, "Model Imprimante", "Modèle Imprimante"))
                claim_date = to_date_iso(first_present(rec, "Date Reclamation", "Date Réclamation"))
                interv_date = to_date_iso(first_present(rec, "Date D'intervention CBI", "Date d'intervention CBI"))
                duration = to_text(first_present(rec, "Duree de traitement ticket", "Duree de traitement ticket "))

                if not (site or printer_name or ticket or nature):
                    continue

                item = models.PrinterTonerIncident(
                    id=gen_id("pti", site, printer_name, ticket, claim_date, nature),
                    site=site or None,
                    printerName=printer_name or None,
                    demandType=demand_type or None,
                    ticketNumber=ticket or None,
                    problemNature=nature or None,
                    printerSerial=serial or None,
                    printerModel=model or None,
                    claimDate=claim_date,
                    interventionDate=interv_date,
                    duration=duration or None,
                    status="INTERVENUE" if interv_date else "NON_INTERVENUE",
                    raw=raw or None,
                    rawHeaders=[str(h).strip() for h in headers if str(h or "").strip()] or None,
                )
                try:
                    DB.printer_toner_incidents.create(item.id, lambda _id, it=item: it)
                except ValueError:
                    continue
    except Exception as e:
        print(f"Incidents restore skipped: {type(e).__name__}: {e}")

    # Entrées
    try:
        ws = wb["Entrées"]
        header = find_header_row(ws, "Date d'entrée")
        if header is not None:
            header_row, headers = header
            for rec in iter_sheet_records(ws, header_row, headers):
                d = to_date_iso(first_present(rec, "Date d'entrée", "Date d entree"))
                article = to_text(first_present(rec, "Article"))
                code = to_text(first_present(rec, "Code Artice", "Code Article"))
                qty = to_int(first_present(rec, "Quantité", "Quantite"))
                if not (article or qty):
                    continue
                item = models.PrinterTonerEntry(
                    id=gen_id("pte", d, article, code, qty),
                    date=d,
                    article=article or None,
                    articleCode=code or None,
                    quantity=int(qty or 0),
                )
                try:
                    DB.printer_toner_entries.create(item.id, lambda _id, it=item: it)
                except ValueError:
                    continue
    except Exception as e:
        print(f"Entries restore skipped: {type(e).__name__}: {e}")

    # Sorties
    try:
        ws = wb["Sorties"]
        header = find_header_row(ws, "Date de sortie")
        if header is not None:
            header_row, headers = header
            for rec in iter_sheet_records(ws, header_row, headers):
                d = to_date_iso(first_present(rec, "Date de sortie"))
                article = to_text(first_present(rec, "Nom Article", "Article"))
                code = to_text(first_present(rec, "Code Artice", "Code Article"))
                qty = to_int(first_present(rec, "Quantité", "Quantite"))
                if not (article or qty):
                    continue
                item = models.PrinterTonerExit(
                    id=gen_id("ptx", d, article, code, qty),
                    date=d,
                    article=article or None,
                    articleCode=code or None,
                    quantity=int(qty or 0),
                )
                try:
                    DB.printer_toner_exits.create(item.id, lambda _id, it=item: it)
                except ValueError:
                    continue
    except Exception as e:
        print(f"Exits restore skipped: {type(e).__name__}: {e}")

    # Min qty
    try:
        if "AS3" in wb.sheetnames:
            ws = wb["AS3"]
            header = find_header_row(ws, "Référence") or find_header_row(ws, "Reference")
            if header is not None:
                header_row, headers = header
                current_ref = ""
                for rec in iter_sheet_records(ws, header_row, headers):
                    ref = to_text(first_present(rec, "Référence ", "Référence", "Reference"))
                    color = to_text(first_present(rec, "Couleur", "Color")).upper()
                    qty = to_int(first_present(rec, "Nombre de toner ", "Nombre de toner", "Nombre de toner  "))
                    if ref:
                        current_ref = ref
                    used_ref = ref or current_ref
                    if not (used_ref and color and qty):
                        continue
                    item = models.PrinterTonerMinQty(
                        id=gen_id("ptm", used_ref, color),
                        ref=used_ref,
                        color=color,
                        minQty=int(qty),
                    )
                    try:
                        DB.printer_toner_min_qty.create(item.id, lambda _id, it=item: it)
                    except ValueError:
                        continue
    except Exception as e:
        print(f"Min qty restore skipped: {type(e).__name__}: {e}")

    wb.close()
    return cleared


def main() -> int:
    backend_dir = Path(__file__).resolve().parents[1]
    repo_root = backend_dir.parent

    env_path = backend_dir / ".env.sqlserver"
    if env_path.exists():
        load_env_file(env_path)

    os.environ["PFE_STORAGE"] = "sqlserver"

    sys.path.insert(0, str(backend_dir))

    from app.sqlserver import connect
    from import_excel_sqlserver import get_table_columns, import_workbook

    cn = connect()
    cur = cn.cursor()

    try:
        from app.storage import DB

        print("Clearing old operational tables...")
        cleared = {
            "assets": clear_repo(DB.assets),
            "stock_movements": clear_repo(DB.movements),
            "assignments": clear_repo(DB.assignments),
            "maintenance_tickets": clear_repo(DB.maintenance_tickets),
            "printer_toner": 0,
        }
        print(cleared)

        print("Restoring printer toner tables...")
        toner_counts = restore_printer_toner_workbook(repo_root)
        print(toner_counts)

        tables = get_table_columns(cur, schema="dbo")

        import_files = [
            repo_root / "database" / "import" / "MA6-Stock Inventory.xlsx",
            repo_root / "database" / "import" / "Inventory-MA6.xlsx",
        ]

        all_stats: Dict[str, Dict[str, int]] = {}
        for path in import_files:
            if not path.exists():
                print(f"Missing workbook: {path}")
                continue
            stats = import_workbook(cur, schema="dbo", path=path, tables=tables, dry_run=False)
            for table, s in stats.items():
                agg = all_stats.setdefault(table, {})
                for k, v in s.items():
                    agg[k] = int(agg.get(k, 0)) + int(v)

        cn.commit()
        print("IMPORT_STATS")
        print(all_stats)

        # Validation counts.
        for table in (
            "assets",
            "stock_movements",
            "assignments",
            "printer_toner_incidents",
            "printer_toner_entries",
            "printer_toner_exits",
            "printer_toner_min_qty",
        ):
            try:
                cur.execute(f"SELECT COUNT(1) FROM dbo.[{table}]")
                print(f"{table}={cur.fetchone()[0]}")
            except Exception as e:
                print(f"{table}=ERR {type(e).__name__}: {e}")

    except Exception as e:
        cn.rollback()
        print(f"Restore failed: {type(e).__name__}: {e}")
        return 1
    finally:
        cur.close()
        cn.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())